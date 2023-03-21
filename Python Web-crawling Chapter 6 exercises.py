import os

'''
print('Current directory :', os.getcwd())
os.chdir("C:\\temp")
print("Moved directory :", os.getcwd())

a = os.listdir("c:\\data")
print(a)

for i in os.listdir("c:\\data") :
    print(i)

print(os.path.exists("c:\\py_temp"))

print(os.path.exists("c:\\data"))


os.mkdir("c://data//test folder")

os.mkdirs("c://data// test folder")

os.rmdir("c://data//test folder")

os.makedirs("c://test//temp 01")

os.removedirs("c://test//temp 01")


os.makedirs("c://py_temp2")
os.chdir("c://py_temp2")
print(os.getcwd())

file = open("test1.txt","w")
file.write("This is the first state in the text file using Python library on Friday 17th March, 2023.")
file.close


file2 = open("test1.txt", "a")
file2.write("\n" + "This is second state in the text file using Python library on Friday 17th March, 2023.")
file2.close()


f = open("C:\\Users\\chois\\Desktop\\Jupyter codes\\Snack samples.txt", "r")
snack = f.readlines()
print(snack)

print("\n")


print(snack[0])
print(snack[2])
print("\n")

for i in snack:
    print(i)
   


import pandas as pd

no = []
subject_name = []

no.append(1)
no.append(2)
no.append(3)

subject_name.append("Math")
subject_name.append("Science")
subject_name.append("Big Data")

subject = pd.DataFrame()
subject['Subject No'] = no
subject['Subject name'] = subject_name
print(subject)

subject.to_csv("C:\\Users\\chois\\Desktop\\Jupyter codes\\Subject 001.csv", encoding = "utf-8-sig", index = False)

subject.to_excel("C:\\Users\\chois\\Desktop\\Jupyter codes\\Subject 001.xlsx", index= False)


import openpyxl

wb = openpyxl.load_workbook("C:\\Users\\chois\\Desktop\\Jupyter codes\\Subject 001.xlsx")
sheet = wb['Sheet1']

contents = {}
for i in range(2, sheet.max_row + 1):
    name = sheet.cell(row=i, column=1).value
    email = sheet.cell(row=i,column=2).value
    contents[name] = email

print(contents)



import csv

f = open("C:\\Users\\chois\\Desktop\\Jupyter codes\\Subject 001.csv")
f_csv = csv.reader(f)
for i in f_csv:
    print(i)


import csv

f = open("C:\\Users\\chois\\Desktop\\Jupyter codes\\Subject 001.csv", encoding="utf-8")
f_csv = csv.reader(f)
for i in f_csv:
    print(i)
    

# Exercise 001

import os

os.chdir("C:\\Users\\chois\\Desktop\\Jupyter codes\\")

f_name = input("Please input your folder name")
exist_f = os.path.exists(f_name)
if exist_f :
    f_name = f_name + "_2"
    os.mkdir(f_name)
else:
    os.mkdir(f_name)
    
'''

import os

f = open("C:\\Users\\chois\\Desktop\\Jupyter codes\\url.txt", "r")
readfile = f.readlines()

i = 0
naver_count = 0
blog_count = 0
tistory_count = 0

for i in range(7):
    if "naver.com" in readfile[i] :
        naver_count += 1
    elif "blog.me" in readfile[i] :
        blog_count += 1
    elif "tistory.com" in readfile[i]:
        tistory_count += 1
    else:
        pass
    
print("blog.naver.com url counted numbers are %s" %naver_count)

print("blog.tistory url counted numbers are %s" %tistory_count)

print("blog.me url counted numbers are %s" %blog_count)
        
